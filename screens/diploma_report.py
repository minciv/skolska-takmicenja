# -*- encoding: utf-8 -*-
# @Аутор      : minciv
# @e-mail     : minciv@protonmail.com
# @Web        : https://github.com/minciv
# @Фајл       : diploma_report.py
# @Верзија    : 0.1.05.
# @Програм    : Windsurf
# @Опис       : Изештај за Вукове и Доситејеве дипломе
# @Датум      : 01.07.2025.

from tkinter import ttk, messagebox, StringVar
from tkinter.constants import *
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

class DiplomaReportFrame(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.parent = parent
        self.app = app
        
        # Иницијализација променљивих
        self.student_var = StringVar()
        self.students_dict = {}  # Мапирање имена и ID-а ученика
        
        # Креирање UI компоненти
        self.setup_ui()
        
        # Учитавање листе ученика
        self.load_students_list()
    
    def setup_ui(self):
        """Креирање корисничког интерфејса"""
        # Главни мени је сада на нивоу целе апликације
        
        # Главни контејнер
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill=BOTH, expand=YES)
        
        # Горњи део - контроле за избор и генерисање извештаја
        control_frame = ttk.LabelFrame(main_frame, text="Параметри извештаја", padding=10)
        control_frame.pack(fill=X, expand=NO, pady=(0, 10))
        
        # Избор ученика
        ttk.Label(control_frame, text="Ученик:").grid(row=0, column=0, sticky=W, pady=5, padx=5)
        self.student_cb = ttk.Combobox(control_frame, textvariable=self.student_var, width=40)
        self.student_cb.grid(row=0, column=1, sticky=W, pady=5, padx=5)
        
        # Дугмад за акције
        button_frame = ttk.Frame(control_frame)
        button_frame.grid(row=0, column=2, padx=20)
        ttk.Button(button_frame, text="Прикажи извештај", command=self.show_report).pack(side=LEFT, padx=5)
        ttk.Button(button_frame, text="Прикажи све ученике 8. разреда", command=self.show_all_8th_grade).pack(side=LEFT, padx=5)
        ttk.Button(button_frame, text="Извези у Excel", command=self.export_to_excel).pack(side=LEFT, padx=5)
        
        # Доњи део - табела са резултатима
        self.result_frame = ttk.LabelFrame(main_frame, text="Резултати", padding=10)
        self.result_frame.pack(fill=BOTH, expand=YES)
        
        # Иницијално празна табела
        self.create_empty_table()
    
    def create_empty_table(self):
        """Креирање празне табеле за резултате"""
        for widget in self.result_frame.winfo_children():
            widget.destroy()
            
        message_label = ttk.Label(
            self.result_frame, 
            text="Изаберите ученика и кликните 'Прикажи извештај' да видите резултате",
            font=('Helvetica', 12)
        )
        message_label.pack(pady=50)
    
    def load_students_list(self):
        """Учитавање листе ученика за combobox"""
        try:
            cursor = self.app.conn.cursor()
            cursor.execute("""
                SELECT student_id, first_name, last_name 
                FROM Students
                WHERE active = 1
                ORDER BY last_name, first_name
            """)
            
            students = []
            for row in cursor.fetchall():
                student_id, first_name, last_name = row
                full_name = f"{first_name} {last_name}"
                students.append(full_name)
                self.students_dict[full_name] = student_id
            
            # Додајемо опцију "Сви ученици 8. разреда"
            all_option = "[ Сви ученици 8. разреда ]"
            students = [all_option] + students
            self.students_dict[all_option] = None
            
            self.student_cb['values'] = students
            
            # Постављамо подразумевану вредност
            if students:
                self.student_var.set(students[0])
                
        except Exception as e:
            messagebox.showerror("Грешка", f"Грешка при учитавању ученика: {str(e)}")
    
    def show_report(self):
        """Приказ извештаја за изабраног ученика"""
        # Прво чистимо претходну табелу
        for widget in self.result_frame.winfo_children():
            widget.destroy()
            
        selected_student = self.student_var.get()
        if not selected_student:
            self.create_empty_table()
            return
        
        # Ако је изабрана опција "Сви ученици 8. разреда"
        if selected_student == "[ Сви ученици 8. разреда ]":
            self.show_all_8th_grade()
            return
        
        student_id = self.students_dict.get(selected_student)
        if not student_id:
            messagebox.showwarning("Упозорење", "Молимо изаберите ученика из листе.")
            return
        
        try:
            cursor = self.app.conn.cursor()
            
            # Проверавамо да ли ученик има успехе
            cursor.execute("""
                SELECT COUNT(*) 
                FROM StudentAchievements 
                WHERE student_id = ?
            """, (student_id,))
            
            if cursor.fetchone()[0] == 0:
                message_label = ttk.Label(
                    self.result_frame, 
                    text=f"Ученик {selected_student} нема евидентираних успеха",
                    font=('Helvetica', 12)
                )
                message_label.pack(pady=50)
                return
            
            # Учитавамо све успехе ученика груписане по разредима
            cursor.execute("""
                SELECT 
                    a.grade,
                    c.competition_name, 
                    l.level_name, 
                    a.placement, 
                    a.school_year,
                    a.mentor_name
                FROM StudentAchievements a
                JOIN Competitions c ON a.competition_id = c.competition_id
                JOIN CompetitionLevels l ON a.level_id = l.level_id
                WHERE a.student_id = ?
                ORDER BY a.grade, c.competition_name
            """, (student_id,))
            
            results = cursor.fetchall()
            
            # Груписање по разредима
            grades_data = {}
            for row in results:
                grade, competition, level, placement, school_year, mentor = row
                if grade not in grades_data:
                    grades_data[grade] = []
                grades_data[grade].append((competition, level, placement, school_year, mentor))
            
            # Креирамо таб контролу за приказ разреда
            notebook = ttk.Notebook(self.result_frame)
            notebook.pack(fill=BOTH, expand=YES)
            
            # Креирамо таб за сваки разред
            for grade in sorted(grades_data.keys()):
                grade_frame = ttk.Frame(notebook, padding=10)
                notebook.add(grade_frame, text=f"{grade}. разред")
                
                # Табела за приказ успеха
                columns = ("Такмичење", "Ниво", "Пласман", "Школска година", "Ментор")
                tree = ttk.Treeview(grade_frame, columns=columns, show="headings")
                tree.pack(fill=BOTH, expand=YES)
                
                # Подешавање заглавља колона
                for col in columns:
                    tree.heading(col, text=col)
                    tree.column(col, width=150)
                
                # Додавање података у табелу
                for achievement in grades_data[grade]:
                    tree.insert("", END, values=achievement)
                
                # Скролбар
                scrollbar = ttk.Scrollbar(grade_frame, orient=VERTICAL, command=tree.yview)
                tree.configure(yscrollcommand=scrollbar.set)
                scrollbar.pack(side=RIGHT, fill=Y)
            
            # Додајемо таб за сумарни приказ
            summary_frame = ttk.Frame(notebook, padding=10)
            notebook.add(summary_frame, text="Сумарно")
            
            # Креирамо табелу са сумарним приказом
            ttk.Label(
                summary_frame, 
                text=f"Сумарни приказ успеха ученика: {selected_student}",
                font=('Helvetica', 12, 'bold')
            ).pack(pady=(0, 10))
            
            summary_tree = ttk.Treeview(
                summary_frame, 
                columns=("Разред", "Број успеха", "Број такмичења", "Најбољи пласман"),
                show="headings"
            )
            summary_tree.pack(fill=X)
            
            for col in ("Разред", "Број успеха", "Број такмичења", "Најбољи пласман"):
                summary_tree.heading(col, text=col)
                summary_tree.column(col, width=150, anchor=CENTER)
            
            total_achievements = 0
            total_competitions = set()
            
            for grade, achievements in sorted(grades_data.items()):
                competitions_set = set([a[0] for a in achievements])
                best_placement = min([a[2] for a in achievements], key=lambda p: 
                                    0 if "1. место" in p else
                                    1 if "2. место" in p else
                                    2 if "3. место" in p else
                                    3 if "Похвала" in p else 4)
                
                summary_tree.insert("", END, values=(
                    f"{grade}. разред",
                    len(achievements),
                    len(competitions_set),
                    best_placement
                ))
                
                total_achievements += len(achievements)
                total_competitions.update(competitions_set)
            
            # Додајемо укупно за све разреде
            summary_tree.insert("", END, values=(
                "УКУПНО",
                total_achievements,
                len(total_competitions),
                ""
            ))
                
        except Exception as e:
            messagebox.showerror("Грешка", f"Грешка при генерисању извештаја: {str(e)}")
    
    def show_all_8th_grade(self):
        """Приказ успеха свих ученика 8. разреда (на основу тренутног датума и године уписа)"""
        # Прво чистимо претходну табелу
        for widget in self.result_frame.winfo_children():
            widget.destroy()
        
        try:
            cursor = self.app.conn.cursor()
            
            # Учитавамо све ученике из базе
            cursor.execute("""
                SELECT 
                    s.student_id,
                    s.first_name || ' ' || s.last_name as student_name,
                    s.grade,
                    s.school_year
                FROM Students s
                WHERE s.active = 1
                ORDER BY s.last_name, s.first_name
            """)
            
            all_students = cursor.fetchall()
            
            # Филтрирамо само ученике који су тренутно у 8. разреду
            # користећи функцију за израчунавање тренутног разреда
            from screens.students_frame import StudentsFrame  # Увоз потребне класе
            student_frame = StudentsFrame(None, self.app)  # Привремена инстанца за приступ методи
            
            students = []
            for student in all_students:
                student_id, student_name, grade, school_year = student
                
                # Провера да ли је ученик тренутно у 8. разреду
                current_grade = student_frame.calculate_current_grade(grade, school_year)
                if current_grade == 8:
                    students.append((student_id, student_name))
            
            if not students:
                message_label = ttk.Label(
                    self.result_frame, 
                    text="Нема евидентираних успеха ученика у 8. разреду",
                    font=('Helvetica', 12)
                )
                message_label.pack(pady=50)
                return
            
            # Креирамо оквир за приказ листе ученика
            ttk.Label(
                self.result_frame, 
                text="Ученици 8. разреда са евидентираним успесима",
                font=('Helvetica', 12, 'bold')
            ).pack(pady=(0, 10))
            
            # Креирамо табелу са ученицима и њиховим успесима
            columns = ("Ученик", "Број успеха", "Број такмичења", "Најбољи пласман")
            tree = ttk.Treeview(self.result_frame, columns=columns, show="headings")
            tree.pack(fill=BOTH, expand=YES)
            
            # Подешавање заглавља колона
            for col in columns:
                tree.heading(col, text=col)
                if col == "Ученик":
                    tree.column(col, width=250)
                else:
                    tree.column(col, width=150, anchor=CENTER)
            
            # За сваког ученика рачунамо статистику успеха током целог школовања
            for student_id, student_name in students:
                cursor.execute("""
                    SELECT 
                        a.placement,
                        c.competition_name
                    FROM StudentAchievements a
                    JOIN Competitions c ON a.competition_id = c.competition_id
                    WHERE a.student_id = ? 
                    ORDER BY a.placement
                """, (student_id,))
                
                achievements = cursor.fetchall()
                competitions_set = set([a[1] for a in achievements])
                
                if achievements:
                    best_placement = min([a[0] for a in achievements], key=lambda p: 
                                       0 if "1. место" in p else
                                       1 if "2. место" in p else
                                       2 if "3. место" in p else
                                       3 if "Похвала" in p else 4)
                else:
                    best_placement = "-"
                
                tree.insert("", END, values=(
                    student_name,
                    len(achievements),
                    len(competitions_set),
                    best_placement
                ))
            
            # Скролбар
            scrollbar = ttk.Scrollbar(self.result_frame, orient=VERTICAL, command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)
            scrollbar.pack(side=RIGHT, fill=Y)
            
        except Exception as e:
            messagebox.showerror("Грешка", f"Грешка при генерисању извештаја: {str(e)}")
    
    def export_to_excel(self):
        """Извоз података о успесима у Excel фајл"""
        selected_student = self.student_var.get()
        if not selected_student:
            messagebox.showwarning("Упозорење", "Молимо изаберите ученика или опцију 'Сви ученици 8. разреда'.")
            return
        
        # Ако је изабрана опција "Сви ученици 8. разреда"
        if selected_student == "[ Сви ученици 8. разреда ]":
            self.export_all_8th_grade()
            return
        
        student_id = self.students_dict.get(selected_student)
        if not student_id:
            messagebox.showwarning("Упозорење", "Молимо изаберите ученика из листе.")
            return
        
        try:
            from tkinter import filedialog
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel фајлови", "*.xlsx")],
                title="Сачувај извештај за ученика"
            )
            
            if not filename:
                return
            
            cursor = self.app.conn.cursor()
            
            # Учитавамо све успехе ученика груписане по разредима
            cursor.execute("""
                SELECT 
                    a.grade,
                    c.competition_name, 
                    l.level_name, 
                    a.placement, 
                    a.school_year,
                    a.mentor_name
                FROM StudentAchievements a
                JOIN Competitions c ON a.competition_id = c.competition_id
                JOIN CompetitionLevels l ON a.level_id = l.level_id
                WHERE a.student_id = ?
                ORDER BY a.grade, c.competition_name
            """, (student_id,))
            
            results = cursor.fetchall()
            
            # Креирање Excel фајла
            wb = Workbook()
            ws = wb.active
            ws.title = "Успеси ученика"
            
            # Додавање заглавља
            ws.append(["Извештај за ученика:", selected_student])
            ws.append([])
            ws.append(["Разред", "Такмичење", "Ниво", "Пласман", "Школска година", "Ментор"])
            
            # Стилови за заглавље
            header_font = Font(bold=True)
            for cell in ws["3:3"]:
                cell.font = header_font
            
            # Груписање по разредима и додавање у Excel
            current_grade = None
            for row in results:
                grade, competition, level, placement, school_year, mentor = row
                
                if grade != current_grade:
                    current_grade = grade
                    # Додајемо празан ред као раздвајач између разреда
                    if ws.max_row > 3:  # Не додајемо пре првог разреда
                        ws.append([])
                
                ws.append([grade, competition, level, placement, school_year, mentor or ""])
            
            # Аутоматско подешавање ширине колона
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Чување фајла
            wb.save(filename)
            messagebox.showinfo("Успех", f"Извештај је успешно извезен у фајл {filename}")
            
        except Exception as e:
            messagebox.showerror("Грешка", f"Грешка при извозу у Excel: {str(e)}")
    
    def export_all_8th_grade(self):
        """Извоз података о успесима свих ученика 8. разреда у Excel"""
        try:
            from tkinter import filedialog
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel фајлови", "*.xlsx")],
                title="Сачувај извештај за ученике 8. разреда"
            )
            
            if not filename:
                return
            
            cursor = self.app.conn.cursor()
            
            # Креирање Excel фајла
            wb = Workbook()
            
            # Креирамо лист за сумарни приказ
            ws_summary = wb.active
            ws_summary.title = "Сумарни приказ"
            
            # Додавање заглавља
            ws_summary.append(["Извештај за ученике 8. разреда"])
            ws_summary.append([])
            ws_summary.append(["Ученик", "Број успеха", "Број такмичења", "Најбољи пласман"])
            
            # Учитавамо све ученике 8. разреда са успесима
            cursor.execute("""
                SELECT DISTINCT
                    s.student_id,
                    s.first_name || ' ' || s.last_name as student_name
                FROM Students s
                JOIN StudentAchievements a ON s.student_id = a.student_id
                WHERE a.grade = 8 AND s.active = 1
                ORDER BY s.last_name, s.first_name
            """)
            
            students = cursor.fetchall()
            
            for student_id, student_name in students:
                # Креирамо нови лист за сваког ученика
                ws_student = wb.create_sheet(title=student_name[:30])  # Ограничавамо дужину имена
                
                # Додајемо заглавље за лист ученика
                ws_student.append(["Успеси ученика:", student_name])
                ws_student.append([])
                ws_student.append(["Такмичење", "Ниво", "Пласман", "Школска година", "Ментор"])
                
                # Стилови за заглавље
                header_font = Font(bold=True)
                for cell in ws_student["3:3"]:
                    cell.font = header_font
                
                # Учитавамо успехе ученика
                cursor.execute("""
                    SELECT 
                        c.competition_name, 
                        l.level_name, 
                        a.placement, 
                        a.school_year,
                        a.mentor_name
                    FROM StudentAchievements a
                    JOIN Competitions c ON a.competition_id = c.competition_id
                    JOIN CompetitionLevels l ON a.level_id = l.level_id
                    WHERE a.student_id = ? AND a.grade = 8
                    ORDER BY c.competition_name
                """, (student_id,))
                
                achievements = cursor.fetchall()
                competitions_set = set([a[0] for a in achievements])
                
                best_placement = min([a[2] for a in achievements], key=lambda p: 
                                   0 if "1. место" in p else
                                   1 if "2. место" in p else
                                   2 if "3. место" in p else
                                   3 if "Похвала" in p else 4)
                
                # Додајемо у сумарни приказ
                ws_summary.append([
                    student_name,
                    len(achievements),
                    len(competitions_set),
                    best_placement
                ])
                
                # Додајемо успехе на лист ученика
                for achievement in achievements:
                    competition, level, placement, school_year, mentor = achievement
                    ws_student.append([competition, level, placement, school_year, mentor or ""])
                
                # Аутоматско подешавање ширине колона
                for column in ws_student.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws_student.column_dimensions[column_letter].width = adjusted_width
            
            # Подешавање ширине колона у сумарном приказу
            for column in ws_summary.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_summary.column_dimensions[column_letter].width = adjusted_width
            
            # Чување фајла
            wb.save(filename)
            messagebox.showinfo("Успех", f"Извештај је успешно извезен у фајл {filename}")
            
        except Exception as e:
            messagebox.showerror("Грешка", f"Грешка при извозу у Excel: {str(e)}")
