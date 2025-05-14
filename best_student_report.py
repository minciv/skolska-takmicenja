# -*- encoding: utf-8 -*-
# @Аутор   : minciv
# @Фајл     : best_student_report.py
# @Верзија  : 0.1.01.
# @Програм  : Windsurf
# @Опис     : Изештај за ђака генерације
# @Датум   : 14.05.2025.

from tkinter import ttk, messagebox, StringVar, IntVar, Canvas
from tkinter.constants import *
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class BestStudentReportFrame(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.parent = parent
        self.app = app
        
        # Иницијализација променљивих
        self.school_year_var = StringVar()
        
        # Креирање UI компоненти
        self.setup_ui()
        
        # Учитавање школских година
        self.load_school_years()
    
    def setup_ui(self):
        """Креирање корисничког интерфејса"""
        # Главни контејнер
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill=BOTH, expand=YES)
        
        # Горњи део - контроле за избор и генерисање извештаја
        control_frame = ttk.LabelFrame(main_frame, text="Параметри извештаја", padding=10)
        control_frame.pack(fill=X, expand=NO, pady=(0, 10))
        
        # Избор школске године
        ttk.Label(control_frame, text="Школска година:").grid(row=0, column=0, sticky=W, pady=5, padx=5)
        self.school_year_cb = ttk.Combobox(control_frame, textvariable=self.school_year_var, width=20)
        self.school_year_cb.grid(row=0, column=1, sticky=W, pady=5, padx=5)
        
        # Дугмад за акције
        button_frame = ttk.Frame(control_frame)
        button_frame.grid(row=0, column=2, padx=20)
        ttk.Button(button_frame, text="Прикажи ранг листу", command=self.show_ranking).pack(side=LEFT, padx=5)
        ttk.Button(button_frame, text="Извези у Excel", command=self.export_to_excel).pack(side=LEFT, padx=5)
        
        # Информативна порука о правилима бодовања
        info_frame = ttk.Frame(control_frame)
        info_frame.grid(row=1, column=0, columnspan=3, sticky=W, pady=5)
        ttk.Label(
            info_frame, 
            text="Напомена: Ранг листа се генерише на основу успеха ученика од 1. до 8. разреда и правила бодовања.",
            font=('Helvetica', 9, 'italic')
        ).pack(side=LEFT)
        
        # Веза до правила бодовања (само за администраторе)
        if self.app.current_user and self.app.current_user["role"] == "administrator":
            rules_link = ttk.Label(
                info_frame,
                text="Правила бодовања",
                font=('Helvetica', 9, 'underline'),
                foreground="blue",
                cursor="hand2"
            )
            rules_link.pack(side=LEFT, padx=5)
            rules_link.bind("<Button-1>", lambda e: self.switch_to_scoring_rules())
        
        # Доњи део - табела са резултатима
        self.result_frame = ttk.LabelFrame(main_frame, text="Ранг листа", padding=10)
        self.result_frame.pack(fill=BOTH, expand=YES)
        
        # Иницијално празна табела
        self.create_empty_table()
    
    def create_empty_table(self):
        """Креирање празне табеле за резултате"""
        for widget in self.result_frame.winfo_children():
            widget.destroy()
            
        message_label = ttk.Label(
            self.result_frame, 
            text="Изаберите школску годину и кликните 'Прикажи ранг листу' да видите резултате",
            font=('Helvetica', 12)
        )
        message_label.pack(pady=50)
    
    def load_school_years(self):
        """Учитавање школских година из базе"""
        try:
            cursor = self.app.conn.cursor()
            cursor.execute("""
                SELECT DISTINCT school_year 
                FROM StudentAchievements
                WHERE grade = 8
                ORDER BY school_year DESC
            """)
            
            years = cursor.fetchall()
            
            # Ако нема уноса, генеришемо тренутну школску годину
            if not years:
                from datetime import datetime
                current_year = datetime.now().year
                years = [(f"{current_year-1}/{current_year}",)]
            
            school_years = [year[0] for year in years]
            self.school_year_cb['values'] = school_years
            
            # Постављамо подразумевану вредност
            if school_years:
                self.school_year_var.set(school_years[0])
                
        except Exception as e:
            messagebox.showerror("Грешка", f"Грешка при учитавању школских година: {str(e)}")
    
    def switch_to_scoring_rules(self):
        """Пребацивање на таб са правилима бодовања"""
        # Пронаћи notebook у ком се налази овај таб
        parent_notebook = self.parent
        parent_notebook.select(2)  # Индекс таба за правила бодовања
    
    def calculate_student_scores(self, school_year):
        """Рачунање бодова ученика на основу правила бодовања"""
        try:
            cursor = self.app.conn.cursor()
            
            # Учитавамо правила бодовања
            cursor.execute("""
                SELECT level_id, placement_description, points 
                FROM ScoringRules
            """)
            
            # Креирамо речник правила бодовања
            scoring_rules = {}
            for rule in cursor.fetchall():
                level_id, placement, points = rule
                if level_id not in scoring_rules:
                    scoring_rules[level_id] = {}
                scoring_rules[level_id][placement] = points
            
            # Ако нема дефинисаних правила, користимо подразумевана
            if not scoring_rules:
                # Добављамо све нивое такмичења
                cursor.execute("SELECT level_id, level_name FROM CompetitionLevels")
                levels = cursor.fetchall()
                
                # Креирамо подразумевана правила
                default_placements = {
                    "1. место": 10,
                    "2. место": 8,
                    "3. место": 6,
                    "Похвала": 4,
                    "Учешће": 2
                }
                
                for level_id, level_name in levels:
                    scoring_rules[level_id] = {}
                    multiplier = level_id  # Множилац на основу нивоа (претпоставка да су нивои: 1=школски, 2=општински, итд.)
                    for placement, base_points in default_placements.items():
                        scoring_rules[level_id][placement] = base_points * multiplier
            
            # Учитавамо ученике 8. разреда - у школској години не мора бити исто као school_year
        # Овде тражимо ученике чија је школска година иста као школска година за коју се генерише извештај
            cursor.execute("""
                SELECT DISTINCT
                    s.student_id,
                    s.first_name || ' ' || s.last_name as student_name,
                    s.class_enrolled
                FROM Students s
                WHERE s.grade = 8 AND s.active = 1
                ORDER BY s.last_name, s.first_name
            """)
            
            students = cursor.fetchall()
            
            # Рачунамо бодове за сваког ученика
            student_scores = []
            
            for student_id, student_name, class_enrolled in students:
                # Учитавамо све успехе ученика од 1. до 8. разреда
                cursor.execute("""
                    SELECT 
                        a.level_id,
                        a.placement,
                        c.competition_name,
                        a.grade,
                        cl.level_name
                    FROM StudentAchievements a
                    JOIN Competitions c ON a.competition_id = c.competition_id
                    JOIN CompetitionLevels cl ON a.level_id = cl.level_id
                    WHERE a.student_id = ? AND a.grade BETWEEN 1 AND 8
                """, (student_id,))
                
                achievements = cursor.fetchall()
                
                # Рачунамо бодове
                total_points = 0
                achievement_details = []
                
                for level_id, placement, competition_name, grade, level_name in achievements:
                    # Ако постоји правило за овај ниво и пласман
                    if level_id in scoring_rules and placement in scoring_rules[level_id]:
                        points = scoring_rules[level_id][placement]
                        total_points += points
                        
                        achievement_details.append({
                            "competition": competition_name,
                            "level": level_name,
                            "grade": grade,
                            "placement": placement,
                            "points": points
                        })
                    else:
                        # Ако нема правила, додељујемо 0 бодова
                        achievement_details.append({
                            "competition": competition_name,
                            "level": level_name,
                            "grade": grade,
                            "placement": placement,
                            "points": 0
                        })
                
                student_scores.append({
                    "student_id": student_id,
                    "student_name": student_name,
                    "class": class_enrolled,
                    "total_points": total_points,
                    "achievements": achievement_details,
                    "achievement_count": len(achievements)
                })
            
            # Сортирамо по броју бодова (опадајуће)
            student_scores.sort(key=lambda x: x["total_points"], reverse=True)
            
            # Додајемо ранг
            for i, student in enumerate(student_scores):
                student["rank"] = i + 1
            
            return student_scores
            
        except Exception as e:
            messagebox.showerror("Грешка", f"Грешка при рачунању бодова: {str(e)}")
            return []
    
    def show_ranking(self):
        """Приказ ранг листе ученика"""
        # Прво чистимо претходну табелу
        for widget in self.result_frame.winfo_children():
            widget.destroy()
            
        school_year = self.school_year_var.get()
        if not school_year:
            messagebox.showwarning("Упозорење", "Молимо изаберите школску годину.")
            return
        
        # Израчунавамо бодове
        student_scores = self.calculate_student_scores(school_year)
        
        if not student_scores:
            info_label = ttk.Label(
                self.result_frame, 
                text=f"Ранг листа ученика 8. разреда за школску годину {school_year} (укључени сви успеси од 1. до 8. разреда)",
                font=('Helvetica', 14, 'bold')
            )
            info_label.pack(pady=(0, 10))
            
            message_label = ttk.Label(
                self.result_frame, 
                text="Нема евидентираних успеха ученика 8. разреда за школску годину",
                font=('Helvetica', 12)
            )
            message_label.pack(pady=50)
            return
        
        # Креирамо контејнер са табелом и детаљима
        table_frame = ttk.Frame(self.result_frame)
        table_frame.pack(fill=BOTH, expand=YES)
        
        # Креирамо таб контролу за приказ ранг листе и детаља
        notebook = ttk.Notebook(table_frame)
        notebook.pack(fill=BOTH, expand=YES)
        
        # Таб за ранг листу
        rank_frame = ttk.Frame(notebook, padding=10)
        notebook.add(rank_frame, text="Ранг листа")
        
        # Таб за детаље бодовања
        details_frame = ttk.Frame(notebook, padding=10)
        notebook.add(details_frame, text="Детаљи бодовања")
        
        # Креирање табеле за ранг листу
        columns = ("Ранг", "Ученик", "Број успеха", "Укупно бодова")
        tree = ttk.Treeview(rank_frame, columns=columns, show="headings")
        tree.pack(fill=BOTH, expand=YES)
        
        # Подешавање заглавља колона
        for col in columns:
            tree.heading(col, text=col)
            if col in ["Ранг", "Број успеха"]:
                tree.column(col, width=100, anchor=CENTER)
            elif col == "Укупно бодова":
                tree.column(col, width=150, anchor=CENTER)
            else:
                tree.column(col, width=250)
        
        # Додавање података у табелу
        for student in student_scores:
            tree.insert("", END, values=(
                student["rank"],
                student["student_name"],
                student["class"],
                student["achievement_count"],
                student["total_points"]
            ))
        
        # Скролбар за ранг листу
        scrollbar = ttk.Scrollbar(rank_frame, orient=VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=RIGHT, fill=Y)
        
        # Креирање табеле за детаље бодовања
        details_label = ttk.Label(
            details_frame, 
            text="Преглед распореда бодовања по ученицима",
            font=('Helvetica', 12, 'bold')
        )
        details_label.pack(pady=(0, 10))
        
        # За сваког ученика креирамо посебну табелу
        for student in student_scores:
            student_frame = ttk.LabelFrame(
                details_frame, 
                text=f"{student['rank']}. {student['student_name']} ({student['class']}) - Укупно бодова: {student['total_points']}",
                padding=5
            )
            student_frame.pack(fill=X, pady=5)
            
            # Табела за успехе
            details_columns = ("Такмичење", "Ниво", "Разред", "Пласман", "Бодови")
            details_tree = ttk.Treeview(student_frame, columns=details_columns, show="headings", height=len(student['achievements']))
            details_tree.pack(fill=X)
            
            # Подешавање заглавља колона
            for col in details_columns:
                details_tree.heading(col, text=col)
                if col == "Бодови":
                    details_tree.column(col, width=100, anchor=CENTER)
                else:
                    details_tree.column(col, width=200)
            
            # Додавање детаља о успесима
            for achievement in student['achievements']:
                details_tree.insert("", END, values=(
                    achievement["competition"],
                    achievement["level"],
                    achievement["grade"],
                    achievement["placement"],
                    achievement["points"]
                ))
        
        # Детаљни приказ успеха изабраног ученика
        details_content = ttk.Frame(details_frame)
        details_content.pack(fill=BOTH, expand=YES, pady=10)
        
        # Заглавље за успехе
        ttk.Label(details_content, text="Такмичење", width=20, font=('Helvetica', 10, 'bold')).grid(row=0, column=0, padx=5, pady=5, sticky=W)
        ttk.Label(details_content, text="Ниво", width=15, font=('Helvetica', 10, 'bold')).grid(row=0, column=1, padx=5, pady=5, sticky=W)
        ttk.Label(details_content, text="Разред", width=10, font=('Helvetica', 10, 'bold')).grid(row=0, column=2, padx=5, pady=5, sticky=W)
        ttk.Label(details_content, text="Пласман", width=15, font=('Helvetica', 10, 'bold')).grid(row=0, column=3, padx=5, pady=5, sticky=W)
        ttk.Label(details_content, text="Бодови", width=10, font=('Helvetica', 10, 'bold')).grid(row=0, column=4, padx=5, pady=5, sticky=W)
        
        # Скролбар за детаље
        details_scrollbar = ttk.Scrollbar(details_frame)
        details_scrollbar.pack(side=RIGHT, fill=Y)
        
        # Canvas за скроловање - користимо Canvas из tkinter не из ttk
        details_canvas = Canvas(details_frame, yscrollcommand=details_scrollbar.set)
        details_canvas.pack(side=LEFT, fill=BOTH, expand=YES)
        details_scrollbar.config(command=details_canvas.yview)
    
    def export_to_excel(self):
        """Извоз ранг листе у Excel фајл"""
        school_year = self.school_year_var.get()
        if not school_year:
            messagebox.showwarning("Упозорење", "Молимо изаберите школску годину.")
            return
        
        # Израчунавамо бодове
        student_scores = self.calculate_student_scores(school_year)
        
        if not student_scores:
            messagebox.showinfo("Информација", f"Нема података за школску годину {school_year}")
            return
        
        try:
            from tkinter import filedialog
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel фајлови", "*.xlsx")],
                title="Сачувај ранг листу за ђака генерације"
            )
            
            if not filename:
                return
            
            # Креирање Excel фајла
            wb = Workbook()
            
            # Креирамо лист за ранг листу
            ws_ranking = wb.active
            ws_ranking.title = "Ранг листа"
            
            # Додавање заглавља
            ws_ranking.append([f"Ранг листа за ђака генерације - школска година {school_year}"])
            ws_ranking.append([])
            ws_ranking.append(["Ранг", "Ученик", "Одељење", "Број успеха", "Укупно бодова"])
            
            # Стил за заглавље
            header_font = Font(bold=True)
            for cell in ws_ranking["3:3"]:
                cell.font = header_font
            
            # Додавање ранг листе
            for student in student_scores:
                ws_ranking.append([
                    student["rank"],
                    student["student_name"],
                    student["class"],
                    student["achievement_count"],
                    student["total_points"]
                ])
            
            # Аутоматско подешавање ширине колона
            for column in ws_ranking.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_ranking.column_dimensions[column_letter].width = adjusted_width
            
            # Креирамо лист за детаље
            ws_details = wb.create_sheet(title="Детаљи бодовања")
            
            # Додавање заглавља
            ws_details.append([f"Детаљи бодовања за ђака генерације - школска година {school_year}"])
            ws_details.append([])
            
            # Додавање детаља за сваког ученика
            row_num = 3
            for student in student_scores:
                # Заглавље ученика
                ws_details.append([
                    f"{student['rank']}. {student['student_name']} ({student['class']}) - Укупно бодова: {student['total_points']}"
                ])
                ws_details.cell(row=row_num, column=1).font = Font(bold=True)
                row_num += 1
                
                # Додавање заглавља
                ws_details.append(["Такмичење", "Ниво", "Разред", "Пласман", "Бодови"])
                for cell in ws_details[row_num]:
                    cell.font = Font(bold=True)
                row_num += 1
                
                # Додавање успеха
                for achievement in student['achievements']:
                    ws_details.append([
                        achievement["competition"],
                        achievement["level"],
                        achievement["grade"],
                        achievement["placement"],
                        achievement["points"]
                    ])
                    row_num += 1
                
                # Додајемо празан ред између ученика
                ws_details.append([])
                row_num += 1
            
            # Аутоматско подешавање ширине колона
            for column in ws_details.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_details.column_dimensions[column_letter].width = adjusted_width
            
            # Чување фајла
            wb.save(filename)
            messagebox.showinfo("Успех", f"Ранг листа је успешно извезена у фајл {filename}")
            
        except Exception as e:
            messagebox.showerror("Грешка", f"Грешка при извозу у Excel: {str(e)}")
